#Olá! Seja bem vindo ao código fonte da votadora
#Urna eletrônica criada para a votação do Grêmio Estudantil da minha escola ( ETE MFL ) no ano de 2024.

#Fazendo as importações necessárias
import customtkinter as ctk
import tkinter
import openpyxl
from datetime import datetime
import pygame
from PIL import Image

#Inicializando o programa
print("--------- Programa Iniciado com sucesso! ---------")

#Criando a janela da aplicação
app = ctk.CTk()
app.wm_title("URNA ETE V0.1")
app.iconbitmap("CPG_ico.ico")
app.geometry("1280x720")
ctk.set_appearance_mode("Light")

print("Janela criada com sucesso!")

#Escrevendo informações na tela
name_label = ctk.CTkLabel(app, text="ELEIÇÕES", font=("Arial", 20, "bold"), bg_color="transparent")
name_label.place(x=10,y=19)

ete_image = ctk.CTkImage(Image.open(r"ete_logo.png"), size=(180,180))
ete_label = ctk.CTkLabel(app, text="", image=ete_image)
ete_label.place(x=5,y=0)

#Desenhando outras áreas, area de informação do voto
info_canvas =  ctk.CTkCanvas(
    app,
    width = 700,
    height = 500,
    bg = "#A5A7B1"
)
info_canvas.place(x=250,y=120)

#Valor padrão do voto
button_value = "?"

#Criando um objeto para utilizar a data e hora
myobj = datetime.now()

#DEFININDO CHAPAS
candidatos = ["FUTURO ESTUDANTIL ( FEST )", "AÇÃO ESTUDANTIL", "TODAS AS VOZES"]

#SOMS DA URNA
pygame.init()
urnasound = pygame.mixer.Sound("confirmsound.mp3")
errorsound = pygame.mixer.Sound("errorsound.mp3")

#CAIXAS ONDE OS DÍGITOS SERÃO VISUALIZADOS
number_entry = ctk.CTkButton(app, text=button_value, corner_radius=10, border_color="white", border_width=0, width=60,height=60, bg_color="#172259", fg_color= "black", font=("Arial",50,"bold"))
number_entry.place(x=300,y=300) 

number_entry2 = ctk.CTkButton(app, text=button_value, corner_radius=10, border_color="white", border_width=0, width=60,height=60, bg_color="#172259", fg_color= "black", font=("Arial",50,"bold"))
number_entry2.place(x=360,y=300)

#------------------ FUNÇÕES --------------------
 
#FUNÇÃO PARA ESCREVER OS DÍGITOS NA CAIXA DE TEXTO 
def write_number(button_name):
    global button_value
    button_value = button_name.cget("text")
    print(number_entry.cget("text"))
    if number_entry.cget("text") != "1" and number_entry.cget("text") != "2" and number_entry.cget("text") != "3" and number_entry.cget("text") != "4" and number_entry.cget("text") != "5" and number_entry.cget("text") != "6" and number_entry.cget("text") != "7" and number_entry.cget("text") != "8" and number_entry.cget("text") != "9" and number_entry.cget("text") != "0":
        number_entry.configure(text=button_value)
    else:
        number_entry2.configure(text=button_value)

#BOTÃO CORRIGIR
def corrige_number():
    number_entry.configure(text="?")
    number_entry2.configure(text="?")

#BOTÃO CONFIRMA_VOTO
def sending_to_sheet():
    global vote_number_now
    number_entry.cget("text")
    number_entry2.cget("text")
    vote_number_now = number_entry.cget("text") + number_entry2.cget("text")
    print(f"O voto inserido pelo usuário atual foi: {vote_number_now}") 
    print("-" * 40)
    print("VOTO COMPUTADO COM SUCESSO!")
    if vote_number_now == "01":
        print(f"A chapa a qual o suário votou foi a de nº{vote_number_now} de nome {candidatos[0]}!, na hora {myobj.hour}:{myobj.minute}:{myobj.second}")
        urnasound.play()
        hour = f"{myobj.hour}:{myobj.minute}:{myobj.second}"
        path = r"C:\Users\Admin\Desktop\Codes\UrnaETE\VOTOS.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        vote_values = [candidatos[0],vote_number_now,hour]
        sheet.append(vote_values)
        workbook.save(path)
    elif vote_number_now == "02":
        print(f"A chapa a qual o suário votou foi a de nº{vote_number_now} de nome {candidatos[1]}!, na hora {myobj.hour}:{myobj.minute}:{myobj.second}")
        urnasound.play()
        hour = f"{myobj.hour}:{myobj.minute}:{myobj.second}"
        path = r"C:\Users\Admin\Desktop\Codes\UrnaETE\VOTOS.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        vote_values = [candidatos[1],vote_number_now,hour]
        sheet.append(vote_values)
        workbook.save(path)
    elif vote_number_now == "03":
        print(f"A chapa a qual o suário votou foi a de nº{vote_number_now} de nome {candidatos[2]}!, na hora {myobj.hour}:{myobj.minute}:{myobj.second}")
        urnasound.play()
        hour = f"{myobj.hour}:{myobj.minute}:{myobj.second}"
        path = r"C:\Users\Admin\Desktop\Codes\UrnaETE\VOTOS.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        vote_values = [candidatos[2],vote_number_now,hour]
        sheet.append(vote_values)
        workbook.save(path)
    else: 
        print("Número de votação inválido. Tente outra vez.")
        errorsound.play()
        response = tkinter.messagebox.askquestion("ESCOLHA UM NÚMERO VÁLIDO!", "DESEJA CONTINUAR VOTANDO?")
        if response == "yes":
            pass
        else:
            app.destroy()
            
#CANVAS DO NÚMERO DOS VOTOS
number_canvas =  ctk.CTkCanvas(
    app,
    width = 300,
    height = 500,
    bg = "#172259"
)
number_canvas.place(x=952,y=120)

#DESENHANDO BOTÕES DE DÍGITOS E AÇÕES
One_button = ctk.CTkButton(app, text="1", corner_radius=10, border_color="white", border_width=0, width=60,height=60, bg_color="#172259", fg_color= "black", font=("Arial",50,"bold"),command=lambda:write_number(One_button))
One_button.place(y=180,x=980)

Two_button = ctk.CTkButton(app, text="2", corner_radius=10, border_color="white", border_width=0, width=60,height=60, bg_color="#172259", fg_color= "black", font=("Arial",50,"bold"),command=lambda:write_number(Two_button))
Two_button.place(y=180,x=1075)

Three_button = ctk.CTkButton(app, text="3", corner_radius=10, border_color="white", border_width=0, width=60,height=60, bg_color="#172259", fg_color= "black", font=("Arial",50,"bold"),command=lambda:write_number(Three_button))
Three_button.place(y=180,x=1175)

Four_button = ctk.CTkButton(app, text="4", corner_radius=10, border_color="white", border_width=0, width=60,height=60, bg_color="#172259", fg_color= "black", font=("Arial",50,"bold"),command=lambda:write_number(Four_button))
Four_button.place(y=265,x=980)

Five_button = ctk.CTkButton(app, text="5", corner_radius=10, border_color="white", border_width=0, width=60,height=60, bg_color="#172259", fg_color= "black", font=("Arial",50,"bold"),command=lambda:write_number(Five_button))
Five_button.place(y=265,x=1075)

Six_button = ctk.CTkButton(app, text="6", corner_radius=10, border_color="white", border_width=0, width=60,height=60, bg_color="#172259", fg_color= "black", font=("Arial",50,"bold"),command=lambda:write_number(Six_button))
Six_button.place(y=265,x=1175)

Seven_button = ctk.CTkButton(app, text="7", corner_radius=10, border_color="white", border_width=0, width=60,height=60, bg_color="#172259", fg_color= "black", font=("Arial",50,"bold"),command=lambda:write_number(Seven_button))
Seven_button.place(y=350,x=980)

Eight_button = ctk.CTkButton(app, text="8", corner_radius=10, border_color="white", border_width=0, width=60,height=60, bg_color="#172259", fg_color= "black", font=("Arial",50,"bold"),command=lambda:write_number(Eight_button))
Eight_button.place(y=350,x=1075)

Nine_button = ctk.CTkButton(app, text="9", corner_radius=10, border_color="white", border_width=0, width=60,height=60, bg_color="#172259", fg_color= "black", font=("Arial",50,"bold"),command=lambda:write_number(Nine_button))
Nine_button.place(y=350,x=1175)

Zero_button = ctk.CTkButton(app, text="0", corner_radius=10, border_color="white", border_width=0, width=60,height=60, bg_color="#172259", fg_color= "black", font=("Arial",50,"bold"),command=lambda:write_number(Zero_button))
Zero_button.place(y=435,x=1075)

Confirm_button = ctk.CTkButton(app, text="Confirma", corner_radius=50, border_color="green", border_width=10, border_spacing= 4, width=140,height=5, bg_color="#172259", fg_color= "green", font=("Arial",25,"bold"), command=lambda:sending_to_sheet())
Confirm_button.place(y=550,x=1035)  

Corrige_button = ctk.CTkButton(app, text="Corrige", corner_radius=50, border_color="orange", border_width=5, width=5,height=5, bg_color="#172259", fg_color= "orange", font=("Arial",25,"bold"),command=lambda:corrige_number())
Corrige_button.place(y=440, x=1145)

Nulo_button = ctk.CTkButton(app, text="Branco", corner_radius=50, border_color="white", border_width=5, width=5,height=5, bg_color="#172259", fg_color= "white", font=("Arial",25,"bold"), text_color="black",)
Nulo_button.place(y=440, x=960)

app.mainloop()
